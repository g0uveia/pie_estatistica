import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

filename = sys.argv[1]

wb = load_workbook(filename + '.xlsx')
ws = wb.active

# itera por coluna

for col in ws.iter_cols(min_col=3):
	gabarito = ''
	for row, cell in enumerate(col):
		if row >= 0:
			if row == 0:
				gabarito = cell.value
			elif row == 1:
				continue
			elif cell.value != gabarito:
				if cell.value != '@' and cell.value != '#':
					cell.value = 'X'

wb.save(filename + '.xlsx')
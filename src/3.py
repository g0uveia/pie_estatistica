###################################################
##    PIEDADE - FORMATA PLANILHAS ESTATISTICA    ##
##-----------------------------------------------##
##    3.PY - AUTOR: AUGUSTO GOUVEIA			   	 ##
##    INFO: Move respostas para única linha      ##
##												 ##
###################################################

import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

filename = sys.argv[1]

wb = load_workbook(filename + '.xlsx')
ws = wb.active

index = 1
count = 0

while (index < ws.max_row + 1):
	row = ws[index]
	print(row[0].value)
	# se primeira coluna estiver preenchida
	if row[0].value:
		# vai para proxima linha
		index += 1

	# se primeira coluna estiver vazia
	else:
		print('index: ' + str(index))
		previous = ws[index-1]
		# desconsidera a posição atual
		count_cells = -2

		# contagem das celulas usadas na ultima linha
		for cell in previous:
			if cell.value:
				count_cells += 1
		print('colunas usadas na anterior: ' + str(count_cells))
		# seleciona todas as celulas usadas dessa linha
		cell_range = 'C' + str(index) + ':' + get_column_letter(count_cells+2) + str(index)

		# move para linha anterior após as células preenchidas
		ws.move_range(cell_range, rows=-1, cols=count_cells)

		# deleta a linha que ficou fazia
		ws.delete_rows(index, 1)

wb.save(filename + '.xlsx')
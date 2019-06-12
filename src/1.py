###################################################
##    PIEDADE - FORMATA PLANILHAS ESTATISTICA    ##
##-----------------------------------------------##
##    1.PY - AUTOR: AUGUSTO GOUVEIA			   	 ##
##    INFO: Exclui dados desnecessarios          ##
##												 ##
###################################################

import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pandas import DataFrame
import pandas as pd

filename = sys.argv[1]

wb = load_workbook(filename + '.xlsx')
ws = wb.active


def find_in (dataset, string):
	for data in dataset:
		if data.value == string:
			return True
	return False


### EXCLUI COLUNAS DESNESCESSARIAS

for index, row in enumerate(ws.iter_cols()):
	# procurar por 'Nota'
	if find_in(row, 'Nota'):
		print('found')
		print(index + 1)

		# excluir colunas que vem depois da primeira ocorrencia
		ws.delete_cols(index + 1, ws.max_column - index)
		break


### SEPARA GABARITO

# Duplica sheet
ws1 = wb.copy_worksheet(ws)
ws1.title = 'GABARITO'

# Exclui linhas abaixo de 8
ws1.delete_rows(9, ws1.max_row - 7)

# Exclui linhas acima de 5
ws1.delete_rows(1, 4)

# Exclui colunas depois de BM
ws1.delete_cols(66, ws1.max_column - 63)

# Exclui colunas antes de AB
ws1.delete_cols(0, 27)

ws1.insert_cols(0, 27)
ws1['A1'] = '%'
ws1['D1'] = 'Gabarito'

### SEPARA DADOS RELEVANTES DAS TURMAS

num_turmas = 0
indexes = []

inicio = 0
final = 0

# Conta numero de turmas e delimita info turmas
for index, row in enumerate(ws.iter_rows()):
	if find_in(row, 'Aluno'):
		num_turmas += 1
		if num_turmas == 1:
			inicio = index-1
		else:
			inicio = index
	if find_in(row, 'Total de alunos listados:'):
		final = index
		indexes.append([inicio, final])

# Iterar pelo numero de turmas
for i in range(num_turmas):
	# Duplicar worksheet
	ws2 = wb.copy_worksheet(ws)
	ws2.title = chr(65+i)
	
	#-- DEBUG --#
	print(wb.worksheets)
	print('indexes: ', indexes[i])
	
	# Excluir excedente
	# Deletar linhas iguais ou maiores a indexes[i][1]
	ws2.delete_rows(indexes[i][1] + 1, ws.max_row - indexes[i][1])
	# Deletar linhas iguais ou menores a indexes[i][0]
	ws2.delete_rows(1, indexes[i][0] + 1)

	# Adicionar numeros que faltam
	# ...
	# Adicionar letra turma
	# ...

# Excluir worksheet com sobras
wb.remove(ws)

# Junta todas as turmas numa só worksheet
sheets = []

for sheet in wb.worksheets:
	sheets.append(DataFrame(sheet.values))

df = pd.concat(sheets)


### SALVA ARQUIVO
df.to_excel(filename + '.xlsx', index=False, header=False)